from flask import jsonify, request, Blueprint, current_app, send_file
from requests.auth import HTTPBasicAuth
from .helpers import extract_info, find_items, find_tema_for_layer, find_source_for_stora
import requests
from xml.etree import ElementTree as ET
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# Create a Blueprint
app = Blueprint("services", __name__, url_prefix="/services")


@app.route("/GS", methods=["GET"])
def get_services_GS():
    result = []

    portal = request.args.get('portal')
    # Debugging: Print the received portal ID
    print(f"Primljeni portal ID: {portal}")

    # Access the configuration from current_app
    portal_info = current_app.config['PORTALS'].get(portal)
    endpoint_url = current_app.config['URL_GS']

    if not portal_info:
        print("Nevažeći portal ID")
        return jsonify({'message': 'Invalid portal ID'}), 400

    BASE_URL = portal_info['GEOSERVER_URL']
    USERNAME = portal_info['USER_GS']
    PASSWORD = portal_info['PASS_GS']
    LAYERS_URL = endpoint_url['LAYERS_URL']

    services_basic_URL = f"{BASE_URL}{LAYERS_URL}"

    services_response = requests.get(services_basic_URL, auth=HTTPBasicAuth(USERNAME, PASSWORD))
    if services_response.status_code != 200:
        print(f"Neuspješno dohvaćanje popisa servisa. Statusni kod: {services_response.status_code}")
        return jsonify(result)

    try:
        services_json = services_response.json().get('layers', {}).get('layer', [])
    except requests.exceptions.JSONDecodeError as e:
        print(f"Error parsing JSON response for services: {e}")
        result.append(
            {'Geoserver Title': 'Not available', 'Geoserver WMS name': service['name'], 'Source': 'Not available',
             'Data': 'Not available'})
        return jsonify(result)

    for service in services_json:
        href_service = service['href']

        href_service = href_service.split('/rest')
        href_service = BASE_URL + href_service[1]

        href_resource_response = requests.get(href_service, auth=HTTPBasicAuth(USERNAME, PASSWORD))

        try:
            href_resource_json = href_resource_response.json()
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for href_service {href_service}: {e}")
            result.append(
                {'Geoserver Title': 'Not available', 'Geoserver WMS name': service['name'], 'Geoserver Stores': 'Not available',
                 'Data': 'Not available'})
            continue

        resourceType = href_resource_json.get('layer', {}).get('resource', {}).get('@class', None)
        resource_URL = href_resource_json.get('layer', {}).get('resource', {}).get('href', None)
        resource_URL = resource_URL.split('/rest')
        resource_URL = BASE_URL + resource_URL[1]

        if not resourceType or not resource_URL:
            print(f"Missing resourceType or resource_URL for href_service {href_service}")
            result.append(
                {'Geoserver Title': 'Not available', 'Geoserver WMS name': service['name'], 'Geoserver Stores': 'Not available',
                 'Data': 'Not available'})
            continue

        services_details_response = requests.get(resource_URL, auth=HTTPBasicAuth(USERNAME, PASSWORD))

        try:
            services_details = services_details_response.json()
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for resource_URL {resource_URL}: {e}")
            result.append({'Geoserver Title': 'Not available', 'Geoserver WMS name': service['name'], 'Geoserver Stores': 'Not available', 'Data': 'Not available'})
            continue

        result.append(extract_info(services_details, resourceType))

    return jsonify(result)


@app.route("/GS/stores", methods=["GET"])
def get_stores_GS():

    portal = request.args.get('portal')
    # Debugging: Print the received portal ID
    print(f"Primljeni portal ID: {portal}")

    # Access the configuration from current_app
    portal_info = current_app.config['PORTALS'].get(portal)
    endpoint_url = current_app.config['URL_GS']

    if not portal_info:
        print("Nevažeći portal ID")
        return jsonify({'message': 'Invalid portal ID'}), 400

    BASE_URL = portal_info['GEOSERVER_URL']
    USERNAME = portal_info['USER_GS']
    PASSWORD = portal_info['PASS_GS']
    WORKSPACES_URL = endpoint_url['WORKSPACES_URL']

    try:
        # List to store results
        results = []

        # Get list of workspaces
        workspaces_url = f"{BASE_URL}{WORKSPACES_URL}"
        response = requests.get(workspaces_url, auth=HTTPBasicAuth(USERNAME, PASSWORD))

        if response.status_code == 200:
            workspaces = response.json().get('workspaces', {}).get('workspace', [])

            for workspace in workspaces:
                workspace_name = workspace['name']

                datastores_url = f"{BASE_URL}/workspaces/{workspace_name}/datastores.json"
                response = requests.get(datastores_url, auth=HTTPBasicAuth(USERNAME, PASSWORD))
                if response.status_code == 200 and response.json()['dataStores']:
                    datastores = response.json().get('dataStores', {}).get('dataStore', [])

                    for datastore in datastores:
                        datastore_name = datastore['name']
                        datastore_url = f"{BASE_URL}/resource/workspaces/{workspace_name}/{datastore_name}/datastore.xml"
                        response = requests.get(datastore_url, auth=HTTPBasicAuth(USERNAME, PASSWORD))
                        host_value = None
                        if response.status_code == 200:
                            featuretypes_xml = response.text
                            root = ET.fromstring(featuretypes_xml) # Parsing XML
                            for entry in root.find('connectionParameters'):
                                try:
                                    if entry.attrib.get('key') == 'host' or entry.attrib.get('key') == 'GeneralizationInfosProviderParam' or entry.attrib.get('key') == 'jndiReferenceName' or entry.attrib.get('key') == 'url':
                                        host_value = entry.text
                                        break
                                except:
                                    host_value = None
                            results.append({datastore_name : host_value})
                        else:
                            print(f"Failed to fetch feature types for datastore {datastore_name} in workspace {workspace_name}. Status code: {response.status_code}")
                elif response.status_code != 200:
                    print(f"Failed to fetch data stores for workspace {workspace_name}. Status code: {response.status_code}")


                coveragestores_url = f"{BASE_URL}/workspaces/{workspace_name}/coveragestores.json"
                response = requests.get(coveragestores_url, auth=HTTPBasicAuth(USERNAME, PASSWORD))
                if response.status_code == 200 and response.json()['coverageStores']:
                    coveragestores = response.json().get('coverageStores', {}).get('coverageStore', [])

                    for coveragestore in coveragestores:
                        coveragestore_name = coveragestore['name']
                        coveragestore_url = f"{BASE_URL}/resource/workspaces/{workspace_name}/{coveragestore_name}/coveragestore.xml"
                        response = requests.get(coveragestore_url, auth=HTTPBasicAuth(USERNAME, PASSWORD))
                        if response.status_code == 200:
                            featuretypes_xml = response.text
                            root = ET.fromstring(featuretypes_xml) # Parsing XM
                            host_value = root.find(".//url")
                            # Extract the URL text
                            if host_value is not None:
                                url = host_value.text
                                host_value = url
                            else:
                                host_value = None
                            results.append({coveragestore_name : host_value})
                        else:
                            print(f"Failed to fetch feature types for datastore {coveragestore_name} in workspace {workspace_name}. Status code: {response.status_code}")
                elif response.status_code != 200:
                    print(f"Failed to fetch coverage stores for workspace {workspace_name}. Status code: {response.status_code}")

        else:
            print(f"Failed to fetch workspaces. Status code: {response.status_code}")

        return results

    except requests.exceptions.RequestException as e:
        print(f"Request error: {e}")
        return []



@app.route("/admin", methods=["GET"])
def get_services_admin():
    result = []

    portal = request.args.get('portal')
    print(f"Primljeni portal ID: {portal}")

    portal_info = current_app.config['PORTALS'].get(portal)
    endpoint_url = current_app.config['URL_ADMIN']

    if not portal_info:
        print("Nevažeći portal ID")
        return jsonify({'message': 'Invalid portal ID'}), 400

    BASE_URL = portal_info['ADMIN_URL']
    USERNAME = portal_info['USER_ADMIN']
    PASSWORD = portal_info['PASS_ADMIN']
    AUTH_URL = endpoint_url['AUTH_URL']
    LAYERS_URL = endpoint_url['LAYERS_URL']

    if BASE_URL:
        # Get AUTH_TOKEN using POST request
        token_URL = f"{BASE_URL}{AUTH_URL}"
        payload = {
            "username": USERNAME,
            "password": PASSWORD
        }
        response = requests.post(token_URL, json=payload)

        if response.status_code != 200:
            print(f"Neuspješno dohvaćanje tokena. Statusni kod: {response.status_code}")
            return jsonify(result)

        try:
            AUTH_TOKEN = response.json().get('token')
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for token: {e}")
            return jsonify(result)

        HEADERS = {'Authorization': f"Bearer {AUTH_TOKEN}"}

        admin_layers_URL = f"{BASE_URL}{LAYERS_URL}"
        admin_response = requests.get(admin_layers_URL, headers=HEADERS)
        if admin_response.status_code != 200:
            print(f"Neuspješno dohvaćanje slojeva s admina. Statusni kod: {admin_response.status_code}")
            return jsonify(result)

        try:
            admin_layers_json = admin_response.json()
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for admin layers: {e}")
            return jsonify(result)

        for layer in admin_layers_json:
            result.append({'layersName' : layer['layerConfig']['label'],
                'layersBodId' : layer['layerConfig']['serverLayerName']})
        return jsonify(result)
    else:
        return jsonify(result)


@app.route("/topics", methods=["GET"])
def get_topics_admin():
    result = []
    teme_list = []  # popis konfigriranih tema na adminu
    teme_layers = []  # za svaku temu dodan popis slojeva koji su sadržani u temi na admiu

    portal = request.args.get('portal')
    print(f"Primljeni portal ID: {portal}")

    portal_info = current_app.config['PORTALS'].get(portal)
    endpoint_url = current_app.config['URL_ADMIN']

    if not portal_info:
        print("Nevažeći portal ID")
        return jsonify({'message': 'Invalid portal ID'}), 400

    BASE_URL = portal_info['ADMIN_URL']
    USERNAME = portal_info['USER_ADMIN']
    PASSWORD = portal_info['PASS_ADMIN']
    AUTH_URL = endpoint_url['AUTH_URL']
    TEME_URL = endpoint_url['TEME_URL']
    TEME_CONFIG_URL = endpoint_url['TEME_CONFIG_URL']

    if BASE_URL:

        # Get AUTH_TOKEN using POST request
        token_URL = f"{BASE_URL}{AUTH_URL}"
        payload = {
            "username": USERNAME,
            "password": PASSWORD
        }
        response = requests.post(token_URL, json=payload)

        if response.status_code != 200:
            print(f"Neuspješno dohvaćanje tokena. Statusni kod: {response.status_code}")
            return jsonify(result)

        try:
            AUTH_TOKEN = response.json().get('token')
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for token: {e}")
            return jsonify(result)

        HEADERS = {'Authorization': f"Bearer {AUTH_TOKEN}"}

        admin_theme_URL = f"{BASE_URL}{TEME_URL}"
        admin_response = requests.get(admin_theme_URL, headers=HEADERS)
        if admin_response.status_code != 200:
            print(f"Neuspješno dohvaćanje tema s admina. Statusni kod: {admin_response.status_code}")
            return jsonify(result)

        try:
            admin_teme_json = admin_response.json()
        except requests.exceptions.JSONDecodeError as e:
            print(f"Error parsing JSON response for admin layers: {e}")
            return jsonify(result)

        for tema in admin_teme_json:
            teme_list.append(tema['mapName'])

        for tema in teme_list:
            URL_tema = f"{BASE_URL}{TEME_CONFIG_URL}{tema}"
            response = requests.get(URL_tema, headers=HEADERS)
            data = response.json()
            layers = list(find_items(data, 'layerBodId'))
            teme_layers.append({tema: layers})

        result = teme_layers

        return jsonify(result)
    else:
        return jsonify(result)


@app.route("/combined", methods=["GET"])
def get_combined_services():
    result = []

    portal = request.args.get('portal')
    gs_response = requests.get(f"{request.host_url}services/GS?portal={portal}").json()
    gs_stores_response = requests.get(f"{request.host_url}services/GS/stores?portal={portal}").json()
    try:
        admin_response = requests.get(f"{request.host_url}services/admin?portal={portal}").json()
        teme_response = requests.get(f"{request.host_url}services/topics?portal={portal}").json()
    except:
        admin_response = None
        teme_response = None

    for item in gs_response:
        if admin_response and item['Geoserver Title'] != 'Not available':
            flag = False  # Reset flag for each item in gs_response
            for layer in admin_response:
                if item['Geoserver WMS name'] == layer['layersBodId']:
                    merged_item = item.copy()
                    merged_item['Geoportal Name'] = layer['layersName']
                    merged_item['Geoportal Topic'] = find_tema_for_layer(teme_response, layer['layersBodId'])
                    merged_item['Source'] = find_source_for_stora(gs_stores_response, item['Geoserver Stores'])
                    result.append(merged_item)
                    flag = True
                    break
            if not flag:
                merged_item = item.copy()
                merged_item['Geoportal Name'] = 'NO CONFIG'
                merged_item['Geoportal Topic'] = 'NO CONFIG' # ako sloj nije konfiguriran na adimnu onda nije ni u temi
                merged_item['Source'] = find_source_for_stora(gs_stores_response, item['Geoserver Stores'])
                result.append(merged_item)
        elif item['Geoserver Title'] == 'Not available':
            merged_item = item.copy()
            merged_item['Geoportal Name'] = 'Not available'
            merged_item['Geoportal Topic'] = 'Not available'
            merged_item['Source'] = 'Not available'
            result.append(merged_item)
        else:
            merged_item = item.copy()
            merged_item['Geoportal Name'] = 'NOT INTEGRATED'
            merged_item['Geoportal Topic'] = 'NOT INTEGRATED'
            merged_item['Source'] = find_source_for_stora(gs_stores_response, item['Geoserver Stores'])
            result.append(merged_item)

    return jsonify(result)


# @app.route('/export', methods=['POST'])
# def export_data():
#     data = request.json.get('data', [])
#     print('Received data for export:', data) # Debugging log
#     df = pd.DataFrame(data)
#     file_path = os.path.expanduser("~/Desktop/geoportal_data.xlsx")
#     df.to_excel(file_path, index=False)
#     return send_file(file_path, as_attachment=True)

@app.route('/export', methods=['POST'])
def export_data():
    data = request.json.get('data', [])
    print('Received data for export:', data)  # Debugging log

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the header
    headers = ["Geoserver Title", "Geoserver WMS name", "Geoportal Name", "Geoportal Topic", "Geoserver Stores", "Source", "Data"]
    ws.append(headers)

    # Define the red fill
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Append the data to the worksheet
    for item in data:
        row = [
            item["Geoserver Title"],
            item["Geoserver WMS name"],
            item["Geoportal Name"],
            item["Geoportal Topic"],
            item["Geoserver Stores"],
            item["Source"],
            item["Data"]
        ]
        ws.append(row)

        # Apply red fill if hasRedBackground is True
        if item.get("hasRedBackground", False):
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    # Save the workbook to a temporary file
    file_path = os.path.expanduser("~/Desktop/geoportal_data.xlsx")
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)